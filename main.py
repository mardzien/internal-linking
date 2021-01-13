from functools import partial
import multiprocessing as mp
import openpyxl
from openpyxl import load_workbook
import time

import pandas as pd
import numpy as np
import spacy
from spacy.matcher import PhraseMatcher

import files
import metrics as m
from scrape import get_information_from_soup


# funkcja zwraca silnik nlp w zależności od stringa (mogą być różne języki lub różne modele dla jednego języka)
def load_nlp_model(name: str):
    """
    Args:
        name: Language model name
    Returns: Language model object
    """
    nlp = spacy.load(name)
    return nlp


# ładowanie URLi z Inputu1 do listy
def load_url_list(filepath: str):
    return files.load_file_to_list(filepath)


# funkcja pomocnicza potrzebna do stworzenie kolumny z podstawą słowotwórczą
def lemmatizer(nlp_model, phrase):
    doc = nlp_model(str(phrase['Słowo kluczowe']))    # rzutowanie na stringa. Wartości liczbowe mogą być wczytywane jako float.
    result = ""
    for token in doc:
        result += f"{token.lemma_} "
    return result[:-1]


# funkcja przerabiająca Input 2 z URL + Słowo kluczowe na URL + podstawa słowotwórcza
def prepare_input_phrases_with_lemmas(nlp_model, filepath: str):
    """
    Loads xlsx file into pandas dataframe and adds lemmas in new column.
    Args:
        nlp_model: Language model
        filepath: Path to xlsx file
    Returns: Pandas dataframe
    """

    phrase_database = pd.read_excel(filepath, engine="openpyxl")
    phrase_database['Lemma'] = phrase_database.apply(partial(lemmatizer, nlp_model), axis=1)
    phrase_database = phrase_database.drop_duplicates(subset=['URL', 'Lemma'])
    phrase_database.to_excel("Lemmas.xlsx")

    print("Stworzono df z frazami lemma.")

    return phrase_database


#Funkcja zapisująca df do arkusza z odpowiednimi parametrami
def write_df_to_excel(filepath, dataframe):
    # parametr 'strings_to_urls': False jest konieczny ze względu na limit URLi w jednym arkuszu ( limit to 65K URLi)
    with pd.ExcelWriter(filepath, options={'strings_to_urls': False}) as writer:
        dataframe.to_excel(writer, sheet_name='Raport')


# funkcja obecnie nie jest już używana
def append_list_to_excel(filename, list_name, sheet_name):
    columns = "BCDE"
    wb = openpyxl.load_workbook(filename)
    sheet = wb.get_sheet_by_name(sheet_name)
    row_number = sheet.max_row
    sheet[f'A{row_number+1}'] = row_number
    for i, column in enumerate(columns):
        sheet[f'{column}{row_number+1}'] = list_name[i]
    wb.save(filename)


# funkcja obecnie nie jest już używana
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.
    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]
    Returns: None
    """

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


# funkcja z mechanizmem dobierania Inputu 2 dla pojedynczego URLa
def get_match_df_from_single_url(nlp_model, matchers, destination_url_list, input_class, url):
    # print("Processing url:", url)
    url_info = get_information_from_soup(url=url, input_class=input_class)
    text = url_info['Tekst']
    number_of_links = url_info['Liczba linków']
    number_of_paragraphs = url_info['Liczba paragrafów']
    href_list = url_info['Tablica linków']

    document = nlp_model(text)
    column_indexes = ['URL źródłowy', 'Liczba linków', 'Liczba paragrafów', 'Link do strony docelowej', 'URL docelowy',
                      'Słowo kluczowe', 'Kontekst']
    df = pd.DataFrame(columns=column_indexes)

    for d_url in destination_url_list:
        links_to_d_url = "Nie"
        for href in href_list:
            if href == d_url:
                links_to_d_url = "Tak"
        if url == d_url:
            print(">>>>>>>>>> Self matched!")
            continue  # przerwanie pętli, żeby nie szukać dopasowań na siebie.

        matcher = matchers[d_url]
        found_matches = matcher(document)

        if len(found_matches) == 0:
            # print("Not found!")
            pass
        else:
            for match_id, start, end in found_matches:  # tuple unpacking - potrzebujemy tylko start oraz end
                phrase = document[start:end]  # fraza pokrewna znaleziona w tekście
                try:
                    span = document[start - 5:end + 6].text  # tworzenie kontekstu dla znalezionej frazy
                except:
                    span = ""

                df = df.append(pd.Series([url, number_of_links, number_of_paragraphs, links_to_d_url, d_url,
                                          phrase.text, span], index=column_indexes), ignore_index=True)

            print(f">>>>>>>>>> Found {len(found_matches)} matches!")
    # Tutaj lock nie powinien wyrządzić dużej szkody.
    # lock
    # append_df_to_excel(filename="Output/report.xlsx", df=df, sheet_name='Raport')
    # lock
    return df


# funkcja inicjująca matchery i zapisująca je w słowniku
# bardzo zasobochłonna i czasochłonna
def init_matchers(nlp_model, database_df, destination_urls):
    d = dict()
    mtr = m.Metrics('matcher')
    for url in destination_urls:
        mtr.start()
        phrase_list = list(database_df['Lemma'][database_df['URL'] == url])
        phrase_patterns = [nlp_model(text) for text in phrase_list]
        matcher = PhraseMatcher(nlp_model.vocab, attr="LEMMA")
        matcher.add(url, None, *phrase_patterns)
        d[url] = matcher
        mtr.stop()
    mtr.report()
    return d


# funkcja wykorzystuje multiprocessing- łączy cały input1 z inputem2, procesy różnią się od siebie URLami z inputu1
def create_inlinks_report(nlp_model, source_url_list, database_df, input_class):
    mp.set_start_method('spawn', True)

    destination_url_list = database_df['URL'].unique()

    print("Initializing matchers")
    matchers = init_matchers(nlp_model, database_df, destination_url_list)
    print("Matchers initialized")

    target = partial(get_match_df_from_single_url, nlp_model, matchers, destination_url_list, input_class)

    # liczba procesów na danej maszynie - 1, ze względu na jeden proces sterujący programem
    proc_num = mp.cpu_count() - 1
    with mp.Pool(proc_num) as p:
        results = p.map(target, source_url_list)

    return results


if __name__ == '__main__':
    mo = m.Metrics('overall')
    mo.start()
    lang_model = load_nlp_model("pl_core_news_sm")
    url_list = load_url_list("Input/url_list2.txt")
    df_phrases = prepare_input_phrases_with_lemmas(lang_model, "Input/Renee_nowe_kategorie.xlsx")

    df_results = create_inlinks_report(lang_model, url_list, df_phrases,
                                       input_class='eltd-post-text-inner')
    df = pd.concat(df_results)
    df.reset_index(drop=True, inplace=True)
    write_df_to_excel("Output/Renee_blog.xlsx", df)

    mo.stop()
    mo.report()

# TODO: Collect metrics and report them
